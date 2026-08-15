"""Build the AB manual-update worklist as an .xlsx.

Consolidates every "someone has to open this and fix it" item found by the audits:
  1. Nested sub-HDAs pinned at an old version inside a producer definition
  2. Nested AB types referenced that no longer exist on disk
  3. Textures referenced via opdef: under an asset name that no longer exists
  4. Orphaned embedded textures (deletable)
  5. Assets with external texture dependencies (break if a sibling is absent)

Read-only. Writes AB_Update_Worklist.xlsx next to this script.
"""
import json, os, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT  = os.path.join(HERE, "AB_Update_Worklist.xlsx")

nested = json.load(open(os.path.join(HERE, "nested_versions.json"), encoding="utf-8"))
graph  = json.load(open(os.path.join(HERE, "opdef_graph.json"), encoding="utf-8"))

HDR  = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="35699E")
SEV  = {"HIGH": PatternFill("solid", fgColor="F4CCCC"),
        "MED":  PatternFill("solid", fgColor="FCE5CD"),
        "LOW":  PatternFill("solid", fgColor="D9EAD3")}

wb = Workbook()
wb.remove(wb.active)


def sheet(name, headers, rows, widths=None):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in ws[1]:
        c.font = HDR; c.fill = FILL; c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(r)
        sev = str(r[0]).upper()
        if sev in SEV:
            for c in ws[ws.max_row]:
                c.fill = SEV[sev]
    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


# ---- 1 + 2. nested versions ------------------------------------------------
stale_rows, missing_rows = [], []
for parent, d in sorted(nested.get("results", {}).items()):
    for s in d.get("stale", []):
        stale_rows.append(["MED", parent, s["nested"], s["newest"],
                           "Open parent, retype nested node to newest, re-save definition"])
    for m in d.get("missing", []):
        missing_rows.append(["HIGH", parent, m,
                             "Type does not exist on disk - was it renamed or removed?"])

sheet("Nested Stale Versions",
      ["Severity", "Parent asset", "Nested sub-asset (pinned)", "Newest available", "Action"],
      stale_rows, [10, 40, 40, 34, 62])
sheet("Nested Missing Types",
      ["Severity", "Parent asset", "Missing nested type", "Action"],
      missing_rows, [10, 40, 42, 60])

# ---- 3. broken opdef texture references ------------------------------------
owned_by = graph["owner_of"]
existing_owners = set()
for a in graph["assets"]:
    base = a[3:].rsplit(".hda", 1)[0]
    parts = base.rsplit(".", 2)
    if len(parts) >= 3:
        existing_owners.add(parts[0])

tex_rows = []
for a, d in sorted(graph["assets"].items()):
    own = set(d["owns"])
    for r in d["refs"]:
        owner, tex = r.split("?", 1)
        short = owner.rsplit("/", 1)[-1].split("::")[0]
        if short in existing_owners:
            continue                      # owner asset still exists - fine
        if tex in own:
            tex_rows.append(["MED", a, tex, short, "Self-owned",
                             "Repoint to extracted shared texture (auto-fixed by extraction)"])
        elif tex in owned_by:
            tex_rows.append(["MED", a, tex, short,
                             "In sibling: " + owned_by[tex][0],
                             "Repoint to extracted shared texture (auto-fixed by extraction)"])
        else:
            tex_rows.append(["HIGH", a, tex, short, "NOT IN LIBRARY",
                             "Texture must be sourced or re-authored"])
sheet("Broken Texture Refs",
      ["Severity", "Asset", "Texture", "References missing asset", "Where the file is", "Action"],
      tex_rows, [10, 38, 38, 26, 34, 58])

# ---- 4. orphans -------------------------------------------------------------
orph_rows = [["LOW", t, ", ".join(owned_by[t]), len(owned_by[t]),
              "Embedded but referenced by nothing - delete rather than extract"]
             for t in sorted(graph["orphan_sections"])]
sheet("Orphan Textures",
      ["Severity", "Texture", "Embedded in", "Copies", "Action"],
      orph_rows, [10, 40, 70, 8, 58])

# ---- 5. duplication ---------------------------------------------------------
dup_rows = [["LOW", t, len(v), ", ".join(sorted(v)[:6]) + (" …" if len(v) > 6 else ""),
             "Extract once to U:/Textures/AB_Embedded/, repoint all copies"]
            for t, v in sorted(graph["duplicated_sections"].items(),
                               key=lambda kv: -len(kv[1]))]
sheet("Duplicated Textures",
      ["Severity", "Texture", "Copies", "Embedded in", "Action"],
      dup_rows, [10, 42, 8, 80, 56])

# ---- summary ----------------------------------------------------------------
ws = wb.create_sheet("Summary", 0)
rows = [
    ("AB Manual-Update Worklist", ""),
    ("generated", "2026-08-14"),
    ("", ""),
    ("Nested sub-HDAs pinned at an old version", len(stale_rows)),
    ("Nested types referenced that do not exist", len(missing_rows)),
    ("Broken texture refs (asset renamed away)", len(tex_rows)),
    ("  of which the file is genuinely absent", sum(1 for r in tex_rows if r[4] == "NOT IN LIBRARY")),
    ("Orphan textures (deletable)", len(orph_rows)),
    ("Textures duplicated across assets", len(dup_rows)),
    ("", ""),
    ("Distinct textures in the library", len(owned_by)),
    ("Total embedded copies", sum(len(v) for v in owned_by.values())),
]
for r in rows:
    ws.append(list(r))
ws["A1"].font = Font(bold=True, size=14)
ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 16

wb.save(OUT)
print("wrote", OUT)
for name in wb.sheetnames:
    print("   %-24s %4d rows" % (name, wb[name].max_row - 1))
